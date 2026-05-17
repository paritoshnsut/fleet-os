"""
ingest.py — Parse client Excel into canonical trip list.
Handles the Adibatala format:
  Col: Sl#, Route Name, Seats, KM/trip, KM/day, In time, Out time, Trip Type
Trip Type = "Both"   → 2 trips (pickup at in-time + drop at out-time)
Trip Type = "Pick Up" → 1 pickup trip
Trip Type = "Drop"    → 1 drop trip
"""

import re
import datetime
import openpyxl
from dataclasses import dataclass, field, asdict
from typing import List, Optional
from config import AVG_SPEED_KMPH


@dataclass
class Trip:
    trip_id:      str
    route_id:     int
    route_name:   str
    trip_type:    str       # 'pickup' | 'drop'
    start_min:    int       # minutes from midnight (e.g. 9:00 → 540)
    end_min:      int       # start_min + duration
    distance_km:  float
    seats_needed: int
    duration_min: int

    def to_dict(self):
        return asdict(self)


def time_to_min(t) -> int:
    """Convert datetime.time or 'HH:MM' string to minutes from midnight."""
    if t is None:
        return -1
    if isinstance(t, datetime.time):
        return t.hour * 60 + t.minute
    if isinstance(t, str):
        m = re.match(r'(\d{1,2})[:\s](\d{2})', t.strip())
        if m:
            return int(m.group(1)) * 60 + int(m.group(2))
    return -1


def duration_from_km(km: float) -> int:
    """Estimate one-way trip duration in minutes based on distance."""
    return max(15, int(km / AVG_SPEED_KMPH * 60))


def normalize_trip_type(raw) -> str:
    if raw is None:
        return ''
    s = str(raw).strip().lower()
    if 'pick' in s:
        return 'pickup'
    if 'drop' in s:
        return 'drop'
    if 'both' in s:
        return 'both'
    return s


def parse_excel(filepath: str) -> tuple[List[Trip], List[str]]:
    """
    Returns (trips, errors).
    trips  — list of Trip objects ready for the scheduler
    errors — list of human-readable validation warnings
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Auto-detect the route sheet (prefer 'Adibatala', else first sheet with the right header)
    sheet = _find_route_sheet(wb)
    if sheet is None:
        return [], ["Could not find a route data sheet. Expected sheet named 'Adibatala' or similar."]

    rows = list(sheet.iter_rows(values_only=True))
    # Find header row (the one that contains 'Route Name')
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return [], ["Could not find header row with 'Route Name' column."]

    col_map = _map_columns(rows[header_idx])

    trips: List[Trip] = []
    errors: List[str] = []
    trip_counter = 0

    for row_i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        # Skip empty rows
        sl = _get(row, col_map, 'sl')
        if sl is None:
            continue

        route_name   = str(_get(row, col_map, 'route_name') or '').strip()
        seats        = int(_get(row, col_map, 'seats') or 36)
        km_per_trip  = float(_get(row, col_map, 'km_per_trip') or 0)
        in_time_raw  = _get(row, col_map, 'in_time')
        out_time_raw = _get(row, col_map, 'out_time')
        trip_type    = normalize_trip_type(_get(row, col_map, 'trip_type'))

        if km_per_trip <= 0:
            errors.append(f"Row {row_i}: Route '{route_name}' has invalid KM ({km_per_trip}), skipped.")
            continue

        in_min  = time_to_min(in_time_raw)
        out_min = time_to_min(out_time_raw)
        dur     = duration_from_km(km_per_trip)
        route_id = int(sl)

        if trip_type in ('pickup', 'both'):
            if in_min < 0:
                errors.append(f"Row {row_i}: Route '{route_name}' has no In-time for pickup, skipped.")
            else:
                trip_counter += 1
                # in_min = office arrival deadline; bus departs homes at (in_min - dur)
                trips.append(Trip(
                    trip_id      = f"P{trip_counter:04d}",
                    route_id     = route_id,
                    route_name   = route_name,
                    trip_type    = 'pickup',
                    start_min    = in_min - dur,
                    end_min      = in_min,
                    distance_km  = km_per_trip,
                    seats_needed = seats,
                    duration_min = dur,
                ))

        if trip_type in ('drop', 'both'):
            if out_min < 0:
                errors.append(f"Row {row_i}: Route '{route_name}' has no Out-time for drop, skipped.")
            else:
                trip_counter += 1
                trips.append(Trip(
                    trip_id      = f"D{trip_counter:04d}",
                    route_id     = route_id,
                    route_name   = route_name,
                    trip_type    = 'drop',
                    start_min    = out_min,
                    end_min      = out_min + dur,
                    distance_km  = km_per_trip,
                    seats_needed = seats,
                    duration_min = dur,
                ))

        if trip_type not in ('pickup', 'drop', 'both'):
            errors.append(f"Row {row_i}: Unknown trip type '{trip_type}' for '{route_name}', skipped.")

    trips.sort(key=lambda t: t.start_min)
    return trips, errors


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _find_route_sheet(wb):
    preferred = ['adibatala', 'adibatla', 'routes', 'route data']
    for name in wb.sheetnames:
        if name.lower() in preferred:
            return wb[name]
    # Fallback: first sheet with 'Route Name' somewhere in first 5 rows
    for name in wb.sheetnames:
        ws = wb[name]
        for row in list(ws.iter_rows(values_only=True))[:5]:
            if any(str(c).lower() == 'route name' for c in row if c):
                return ws
    return None


def _find_header_row(rows):
    for i, row in enumerate(rows[:10]):
        if any(str(c).lower() == 'route name' for c in row if c):
            return i
    return None


def _map_columns(header_row):
    mapping = {}
    keywords = {
        'sl':          ['sl #', 'sl#', 'sl no', 'serial'],
        'route_name':  ['route name'],
        'seats':       ['no. of seats', 'seats', 'no of seats'],
        'km_per_trip': ['km per trip', 'kms per trip'],
        'km_per_day':  ['km per day', 'kms per day'],
        'in_time':     ['in time'],
        'out_time':    ['out time'],
        'trip_type':   ['trip type', 'pickup/ drop', 'pickup/drop'],
    }
    for col_i, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_lower = str(cell).lower().strip()
        for key, variants in keywords.items():
            if any(v in cell_lower for v in variants):
                if key not in mapping:
                    mapping[key] = col_i
    return mapping


def _get(row, col_map, key):
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def trips_summary(trips: List[Trip]) -> dict:
    pickup = [t for t in trips if t.trip_type == 'pickup']
    drop   = [t for t in trips if t.trip_type == 'drop']
    return {
        "total_trips":    len(trips),
        "pickup_trips":   len(pickup),
        "drop_trips":     len(drop),
        "total_route_km": round(sum(t.distance_km for t in trips), 1),
        "unique_routes":  len(set(t.route_id for t in trips)),
        "seat_classes":   sorted(set(t.seats_needed for t in trips)),
        "in_times":       sorted(set(t.end_min for t in pickup)),
        "out_times":      sorted(set(t.start_min for t in drop)),
    }
